import openpyxl, glob, re, json

def clean_text(s):
    if s is None:
        return None
    return re.sub(r'\s+', ' ', str(s)).strip()

def find_header_info(ws):
    header_row = None
    row_vals = None
    for r in range(1, 8):
        vals = [c.value for c in ws[r]]
        norm = [str(v).strip().lower() if isinstance(v, str) else v for v in vals]
        if 'pin' in norm or 'name' in norm:
            header_row = r
            row_vals = vals
            break
    if header_row is None:
        return None
    name_col = desig_col = pin_col = fixedid_col = None
    for idx, v in enumerate(row_vals):
        if isinstance(v, str):
            vv = v.strip().lower()
            if vv == 'pin':
                pin_col = idx
            elif vv in ('fixed id', 'id'):
                fixedid_col = idx
            elif vv == 'name':
                name_col = idx
            elif vv == 'designation':
                desig_col = idx
    return header_row, name_col, desig_col, pin_col, fixedid_col


def norm_pin(pin_val, fixedid_val):
    # Prefer fixed id (zero-padded string) if it looks valid
    if fixedid_val is not None:
        s = str(fixedid_val).strip()
        if s.isdigit():
            return s.zfill(10)
    if pin_val is not None:
        try:
            n = int(pin_val)
            return str(n).zfill(10)
        except (ValueError, TypeError):
            pass
    return None


def parse_month_file(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = "All" if "All" in wb.sheetnames else ("ALL" if "ALL" in wb.sheetnames else None)
    if sheet_name is None:
        raise ValueError(f"No All/ALL sheet in {filepath}")
    ws = wb[sheet_name]
    info = find_header_info(ws)
    if info is None:
        raise ValueError(f"Could not find header row in {filepath}")
    header_row, name_col, desig_col, pin_col, fixedid_col = info

    records = {}  # pin -> {name, designation, zone, area, branch}
    cur_zone = cur_area = cur_branch = None

    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue
        col0 = row[0] if len(row) > 0 else None
        name_val = row[name_col] if name_col is not None and name_col < len(row) else None

        is_data_row = isinstance(col0, (int, float)) and name_val not in (None, '')

        if is_data_row:
            pin_val = row[pin_col] if pin_col is not None and pin_col < len(row) else None
            fixedid_val = row[fixedid_col] if fixedid_col is not None and fixedid_col < len(row) else None
            pin = norm_pin(pin_val, fixedid_val)
            if pin is None:
                continue
            designation = row[desig_col] if desig_col is not None and desig_col < len(row) else None
            name = str(name_val).strip() if name_val else None
            records[pin] = {
                'name': name,
                'designation': clean_text(designation),
                'zone': cur_zone,
                'area': cur_area,
                'branch': cur_branch,
            }
        elif isinstance(col0, str) and col0.strip():
            text = clean_text(col0)
            low = text.lower()
            if low.endswith('zone') or low.endswith('region'):
                cur_zone = text
                cur_area = None
                cur_branch = None
            elif low.endswith('area'):
                cur_area = text
                cur_branch = None
            elif low.endswith('branch'):
                cur_branch = text
            # else: ignore (titles, totals, etc.)

    return records


if __name__ == '__main__':
    files = sorted(glob.glob("25-26 Fiscal Year/*.xlsx"))
    for f in files:
        recs = parse_month_file(f)
        print(f, "-> employees:", len(recs))
        # print a sample
        sample_key = list(recs.keys())[0]
        print("   sample:", sample_key, recs[sample_key])


def extract_hierarchy(filepath):
    """Scan a month file's All/ALL sheet and return the branch->area->zone
    hierarchy as it appears in the section headers, regardless of whether
    staff are currently posted there."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = "All" if "All" in wb.sheetnames else ("ALL" if "ALL" in wb.sheetnames else None)
    ws = wb[sheet_name]
    cur_zone = cur_area = None
    # zone -> area -> set(branches); area with no branch tracked separately
    tree = {}
    area_has_branch = {}

    for row in ws.iter_rows(values_only=True):
        col0 = row[0] if len(row) > 0 else None
        if isinstance(col0, str) and col0.strip():
            text = clean_text(col0)
            low = text.lower()
            if low.endswith('zone') or low.endswith('region'):
                cur_zone = text
                cur_area = None
                tree.setdefault(cur_zone, {})
            elif low.endswith('area'):
                cur_area = text
                if cur_zone is not None:
                    tree.setdefault(cur_zone, {}).setdefault(cur_area, set())
                    area_has_branch.setdefault((cur_zone, cur_area), False)
            elif low.endswith('branch'):
                if cur_zone is not None and cur_area is not None:
                    tree[cur_zone][cur_area].add(text)
                    area_has_branch[(cur_zone, cur_area)] = True

    # convert sets to sorted lists
    result = {}
    for zone, areas in tree.items():
        result[zone] = {area: sorted(branches) for area, branches in areas.items()}
    return result
